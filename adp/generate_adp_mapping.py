import re
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd
import pymysql


# Paths are relative to the script location (this script lives in `imes/adp/`).
# The source ADP files are two levels up under `../../adp/`.
ADP_PATH = Path("../../adp/KISUMU ADP 2025-2026-1-250-321_UPDATED.xlsx")
TEMPLATE_PATH = Path("../../adp/adp_mapping_template.xls")
OUTPUT_PATH = Path("../../adp/adp_mapping_output.xlsx")

# Database connection settings
# Try kisumu_db first, fallback to imbesdb
DB_CONFIG = {
    'host': 'localhost',
    'port': 3307,
    'user': 'impesUser',
    'password': 'Admin2010impes',
    'database': 'imbesdb',  # Using imbesdb as it contains the data
    'charset': 'utf8mb4'
}


def load_template_columns():
    """
    Read the template just to get the expected column order.
    """
    xls = pd.ExcelFile(TEMPLATE_PATH)
    sheet_name = xls.sheet_names[0]
    df = xls.parse(sheet_name, nrows=0)
    return list(df.columns)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    return str(value).strip()


def detect_program_and_department(sheet):
    """
    Heuristic:
    - Look for a cell starting with 'Programme Name' -> Program
    - Look for a cell containing the ADP title text
      e.g. 'Capital and non-Capital projects for the FY 2025/26-<DEPARTMENT>'
      and use ONLY the text after the FY/year part (and following hyphen) as Department.
    """
    program = ""
    department = ""

    for row in range(1, 10):
        for col in range(1, 8):
            value = clean_text(sheet.cell(row=row, column=col).value)
            low = value.lower()
            # Normalized (collapsed) whitespace version for easier matching
            low_norm = re.sub(r"\s+", " ", low)

            if not program and low.startswith("programme name"):
                # After colon is the program name
                parts = value.split(":", 1)
                program = parts[1].strip() if len(parts) > 1 else value.strip()

            # Department detection:
            # Use ONLY the words after "Capital and non-Capital projects for the FY 2025/26-"
            # (or similar variants like 2025/2026). Anything before that is ignored.
            if not department and "capital and non" in low_norm and "projects" in low_norm and "fy" in low_norm:
                # Try to locate the "FY <year/year>" portion - handle both 2025/26 and 2025/2026 formats
                m = re.search(r"fy\s*20\d{2}\s*/\s*\d{2,4}", low, flags=re.IGNORECASE)
                if m:
                    # Take everything after the FY part from the ORIGINAL text
                    after_fy = value[m.end():]
                else:
                    # Fallback: use the whole cell text
                    after_fy = value

                # Often the pattern is "... FY 2025/26- Department Name" or "... FY 2025/2026- Department Name"
                # so we take the text after the first '-' following the FY section.
                if "-" in after_fy:
                    department = after_fy.split("-", 1)[-1].strip()
                else:
                    department = after_fy.strip()
                
                # Clean up any trailing punctuation or extra spaces
                department = department.strip('.,;:')

    return department, program


def detect_header_row(sheet):
    """
    Find the row that contains 'Project name and Location' (or similar),
    which should be the table header.
    """
    for row in range(1, 15):
        values = [clean_text(sheet.cell(row=row, column=c).value) for c in range(1, 10)]
        row_text = " | ".join(values).lower()
        if "project name" in row_text and "location" in row_text:
            return row, values
    return None, None


def build_column_index_map(header_values):
    """
    Map from logical template column names to column indices in the ADP sheet.
    """
    col_map = {}
    for idx, raw in enumerate(header_values):
        name = clean_text(raw).lower()
        col = idx + 1
        if "project" in name and "location" in name:
            col_map["Project name and Location (Ward/Sub county/ county wide)"] = col
        elif "description" in name and "activities" in name:
            col_map["Description of activities"] = col
        elif "green" in name and "economy" in name:
            col_map["Green Economy consideration"] = col
        elif "estimated" in name and "cost" in name:
            col_map["Estimated Cost"] = col
        elif "source" in name and "fund" in name:
            col_map["Source of funds"] = col
        elif "time" in name and "frame" in name:
            col_map["Time frame"] = col
        elif "target" in name:
            col_map["Targets"] = col
    return col_map


def extract_year_from_timeframe(timeframe: str) -> str:
    """
    Extract a compact year/timeframe token from values like '2025-2026', '2023-\\n2027', etc.
    """
    if not timeframe:
        return "2025_26"
    tf = timeframe.replace("\n", " ")
    years = re.findall(r"(20\d{2})", tf)
    if not years:
        return tf.replace(" ", "")
    if len(years) == 1:
        return years[0]
    return f"{years[0][-2:]}{years[1][-2:]}"


def make_project_ref(project_name: str, program: str, timeframe: str, counter: int) -> str:
    """
    Generate a semi-human-readable unique reference:
    {PROG}-{PROJ}-{YR}-{NNN}
    where:
      PROG = initials of program words (max 3)
      PROJ = first 4 letters of main word in project name
      YR   = compact year token from timeframe
      NNN  = incremental counter, 3 digits
    """
    def initials(text: str, max_letters: int = 3) -> str:
        words = re.findall(r"[A-Za-z]+", text)
        return "".join(w[0].upper() for w in words[:max_letters]) or "PRG"

    def proj_token(text: str, length: int = 4) -> str:
        words = re.findall(r"[A-Za-z]+", text)
        if not words:
            return "PRJ"
        return words[0][:length].upper()

    prog_part = initials(program)
    proj_part = proj_token(project_name)
    year_part = extract_year_from_timeframe(timeframe)
    num_part = f"{counter:03d}"
    return f"{prog_part}-{proj_part}-{year_part}-{num_part}"


def expand_shortcuts(text: str) -> str:
    """
    Expand common shortcuts in location names.
    Examples:
    - SW -> SOUTH WEST
    - SE -> SOUTH EAST
    - NW -> NORTH WEST
    - NE -> NORTH EAST
    - N -> NORTH
    - S -> SOUTH
    - E -> EAST
    - W -> WEST
    """
    if not text:
        return text
    
    # Dictionary of shortcuts to full names
    shortcuts = {
        r'\bSW\b': 'SOUTH WEST',
        r'\bSE\b': 'SOUTH EAST',
        r'\bNW\b': 'NORTH WEST',
        r'\bNE\b': 'NORTH EAST',
        r'\bN\b(?!\w)': 'NORTH',  # N not followed by word char
        r'\bS\b(?!\w)': 'SOUTH',  # S not followed by word char
        r'\bE\b(?!\w)': 'EAST',   # E not followed by word char
        r'\bW\b(?!\w)': 'WEST',   # W not followed by word char
    }
    
    expanded = text.upper()
    for pattern, replacement in shortcuts.items():
        expanded = re.sub(pattern, replacement, expanded, flags=re.IGNORECASE)
    
    return expanded


def normalize_for_matching(text: str) -> str:
    """
    Normalize text for case-insensitive matching.
    Removes extra spaces, converts to uppercase, and handles common variations.
    Also handles camelCase by inserting spaces before capital letters.
    """
    if not text:
        return ""
    # Convert camelCase to space-separated (e.g., "countyWide" -> "county Wide")
    text = re.sub(r'([a-z])([A-Z])', r'\1 \2', str(text))
    # Convert to uppercase, remove extra spaces, remove special chars for comparison
    normalized = re.sub(r'[^\w\s]', ' ', text.upper())
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized


def connect_to_database():
    """
    Connect to the database. Try Docker first, then direct connection.
    """
    try:
        # Try direct connection first (if port is exposed)
        connection = pymysql.connect(**DB_CONFIG)
        print("Connected directly to database.")
        return connection
    except Exception as e:
        print(f"Direct connection failed: {e}")
        # Try via Docker exec
        try:
            # Check if Docker container is running
            result = subprocess.run(
                ['docker', 'ps', '--filter', 'name=kisumu_db', '--format', '{{.Names}}'],
                capture_output=True,
                text=True
            )
            if 'kisumu_db' in result.stdout:
                print("Docker container found. Using Docker exec for queries.")
                return None  # Will use Docker exec for queries
            else:
                # Try to find any running mysql container
                result_all = subprocess.run(
                    ['docker', 'ps', '--filter', 'ancestor=mysql:8.0', '--format', '{{.Names}}'],
                    capture_output=True,
                    text=True
                )
                if result_all.stdout.strip():
                    print(f"Found MySQL container: {result_all.stdout.strip()}. Trying to use it.")
                    # Update DB_CONFIG to use the found container
                    return None  # Will use Docker exec
                raise Exception("Docker container not running")
        except Exception as e2:
            print(f"Docker check failed: {e2}")
            # Try alternative: maybe database is accessible via different host
            try:
                alt_config = DB_CONFIG.copy()
                alt_config['host'] = '127.0.0.1'
                connection = pymysql.connect(**alt_config)
                print("Connected to database via 127.0.0.1")
                return connection
            except:
                raise Exception("Could not connect to database. Please ensure Docker container is running.")


def fetch_subcounties_and_wards(connection=None):
    """
    Fetch all subcounties and wards from the database, including ward-to-subcounty relationships.
    Returns a tuple of (subcounties_dict, wards_dict, ward_to_subcounty_dict) where:
    - subcounties_dict: keys are normalized names, values are original names
    - wards_dict: keys are normalized names, values are original names
    - ward_to_subcounty_dict: keys are original ward names, values are original subcounty names
    """
    subcounties = {}
    wards = {}
    ward_to_subcounty = {}  # Maps original ward name to original subcounty name
    
    if connection:
        # Direct connection
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT name FROM kemri_subcounties WHERE voided = 0 OR voided IS NULL")
                for row in cursor.fetchall():
                    name = row[0]
                    normalized = normalize_for_matching(name)
                    if normalized:
                        subcounties[normalized] = name  # Store original name
                
                # Fetch wards with their subcounty relationships
                cursor.execute("""
                    SELECT w.name as ward_name, s.name as subcounty_name 
                    FROM kemri_wards w
                    INNER JOIN kemri_subcounties s ON w.subcountyId = s.subcountyId
                    WHERE (w.voided = 0 OR w.voided IS NULL) 
                    AND (s.voided = 0 OR s.voided IS NULL)
                """)
                for row in cursor.fetchall():
                    ward_name = row[0]
                    subcounty_name = row[1]
                    normalized_ward = normalize_for_matching(ward_name)
                    if normalized_ward:
                        wards[normalized_ward] = ward_name  # Store original name
                        ward_to_subcounty[ward_name] = subcounty_name  # Map ward to subcounty
        except Exception as e:
            print(f"Error fetching from database: {e}")
            # Try with root user
            try:
                root_config = DB_CONFIG.copy()
                root_config['user'] = 'root'
                root_config['password'] = 'root_password'
                root_conn = pymysql.connect(**root_config)
                with root_conn.cursor() as cursor:
                    cursor.execute("SELECT name FROM kemri_subcounties WHERE voided = 0 OR voided IS NULL")
                    for row in cursor.fetchall():
                        name = row[0]
                        normalized = normalize_for_matching(name)
                        if normalized:
                            subcounties[normalized] = name
                    
                    # Fetch wards with their subcounty relationships
                    cursor.execute("""
                        SELECT w.name as ward_name, s.name as subcounty_name 
                        FROM kemri_wards w
                        INNER JOIN kemri_subcounties s ON w.subcountyId = s.subcountyId
                        WHERE (w.voided = 0 OR w.voided IS NULL) 
                        AND (s.voided = 0 OR s.voided IS NULL)
                    """)
                    for row in cursor.fetchall():
                        ward_name = row[0]
                        subcounty_name = row[1]
                        normalized_ward = normalize_for_matching(ward_name)
                        if normalized_ward:
                            wards[normalized_ward] = ward_name
                            ward_to_subcounty[ward_name] = subcounty_name
                root_conn.close()
                print("Successfully fetched data using root user.")
            except Exception as e2:
                print(f"Error fetching with root user: {e2}")
    else:
        # Use Docker exec - try to find the right container name
        container_name = 'kisumu_db'
        try:
            # First check if kisumu_db exists
            result_check = subprocess.run(
                ['docker', 'ps', '--filter', 'name=kisumu_db', '--format', '{{.Names}}'],
                capture_output=True,
                text=True
            )
            if 'kisumu_db' not in result_check.stdout:
                # Try to find any mysql container
                result_all = subprocess.run(
                    ['docker', 'ps', '--filter', 'ancestor=mysql:8.0', '--format', '{{.Names}}'],
                    capture_output=True,
                    text=True
                )
                if result_all.stdout.strip():
                    container_name = result_all.stdout.strip().split('\n')[0]
                    print(f"Using container: {container_name}")
            
            # Try with impesUser first
            result = subprocess.run(
                ['docker', 'exec', container_name, 'mysql', '-u', DB_CONFIG['user'], 
                 f'-p{DB_CONFIG["password"]}', DB_CONFIG['database'], 
                 '-e', "SELECT name FROM kemri_subcounties WHERE voided = 0 OR voided IS NULL"],
                capture_output=True,
                text=True
            )
            if result.returncode != 0:
                # Try with root user
                print("Trying with root user...")
                result = subprocess.run(
                    ['docker', 'exec', container_name, 'mysql', '-u', 'root', 
                     '-proot_password', DB_CONFIG['database'], 
                     '-e', "SELECT name FROM kemri_subcounties WHERE voided = 0 OR voided IS NULL"],
                    capture_output=True,
                    text=True
                )
            
            if result.returncode == 0:
                lines = result.stdout.strip().split('\n')[1:]  # Skip header
                for line in lines:
                    if line.strip() and not line.startswith('Warning'):
                        name = line.strip()
                        normalized = normalize_for_matching(name)
                        if normalized:
                            subcounties[normalized] = name
            else:
                print(f"Error fetching subcounties: {result.stderr}")
            
            # Fetch wards with their subcounty relationships
            query = """SELECT w.name as ward_name, s.name as subcounty_name 
                       FROM kemri_wards w
                       INNER JOIN kemri_subcounties s ON w.subcountyId = s.subcountyId
                       WHERE (w.voided = 0 OR w.voided IS NULL) 
                       AND (s.voided = 0 OR s.voided IS NULL)"""
            
            result = subprocess.run(
                ['docker', 'exec', container_name, 'mysql', '-u', DB_CONFIG['user'], 
                 f'-p{DB_CONFIG["password"]}', DB_CONFIG['database'], 
                 '-e', query],
                capture_output=True,
                text=True
            )
            if result.returncode != 0:
                # Try with root user
                result = subprocess.run(
                    ['docker', 'exec', container_name, 'mysql', '-u', 'root', 
                     '-proot_password', DB_CONFIG['database'], 
                     '-e', query],
                    capture_output=True,
                    text=True
                )
            
            if result.returncode == 0:
                lines = result.stdout.strip().split('\n')[1:]  # Skip header
                for line in lines:
                    if line.strip() and not line.startswith('Warning'):
                        # Parse the two columns: ward_name and subcounty_name
                        parts = line.strip().split('\t')
                        if len(parts) >= 2:
                            ward_name = parts[0].strip()
                            subcounty_name = parts[1].strip()
                            normalized_ward = normalize_for_matching(ward_name)
                            if normalized_ward:
                                wards[normalized_ward] = ward_name
                                ward_to_subcounty[ward_name] = subcounty_name
            else:
                print(f"Error fetching wards: {result.stderr}")
        except Exception as e:
            print(f"Error fetching from database via Docker: {e}")
    
    return subcounties, wards, ward_to_subcounty


def fetch_departments(connection=None):
    """
    Fetch all departments from the database.
    Returns a dict where keys are normalized names and values are original names.
    """
    departments = {}
    
    if connection:
        # Direct connection
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT name FROM kemri_departments WHERE voided = 0 OR voided IS NULL")
                for row in cursor.fetchall():
                    name = row[0]
                    normalized = normalize_for_matching(name)
                    if normalized:
                        departments[normalized] = name  # Store original name
        except Exception as e:
            print(f"Error fetching departments from database: {e}")
    else:
        # Use Docker exec
        try:
            result = subprocess.run(
                ['docker', 'exec', 'kisumu_db', 'mysql', '-u', DB_CONFIG['user'], 
                 f'-p{DB_CONFIG["password"]}', DB_CONFIG['database'], 
                 '-e', "SELECT name FROM kemri_departments WHERE voided = 0 OR voided IS NULL"],
                capture_output=True,
                text=True
            )
            if result.returncode == 0:
                lines = result.stdout.strip().split('\n')[1:]  # Skip header
                for line in lines:
                    if line.strip():
                        name = line.strip()
                        normalized = normalize_for_matching(name)
                        if normalized:
                            departments[normalized] = name
        except Exception as e:
            print(f"Error fetching departments from database via Docker: {e}")
    
    return departments


def match_department_name(department_name: str, departments: dict) -> str:
    """
    Try to match department name from the detected department with database departments.
    Uses word-based matching to handle variations like:
    "Agriculture, Fisheries, Livestock Development & Irrigation" 
    matching "Agriculture, Irrigation, Livestock Fisheries and Blue Economy"
    Returns the department name as found in database, or 'unknown' if not found.
    """
    if not department_name:
        return 'unknown'
    
    normalized_dept = normalize_for_matching(department_name)
    if not normalized_dept:
        return 'unknown'
    
    # Extract key words from the department name (remove common words)
    dept_words = set(re.findall(r'\b\w{3,}\b', normalized_dept))  # Words with 3+ chars
    common_words = {'and', 'the', 'for', 'development', 'department', 'of', 'blue', 'economy'}
    dept_words = dept_words - common_words
    
    # Try exact match first
    if normalized_dept in departments:
        return departments[normalized_dept]
    
    # Try partial matching (check if department name contains or is contained in any database department)
    for norm_name, orig_name in departments.items():
        if normalized_dept in norm_name or norm_name in normalized_dept:
            return orig_name
    
    # Try word-based matching - count how many key words match
    best_match = None
    best_score = 0
    
    for norm_name, orig_name in departments.items():
        # Extract key words from database department name
        db_words = set(re.findall(r'\b\w{3,}\b', norm_name))
        db_words = db_words - common_words
        
        # Count matching words
        matching_words = dept_words & db_words
        if matching_words:
            # Calculate score: number of matching words / total unique words
            total_unique_words = len(dept_words | db_words)
            score = len(matching_words) / total_unique_words if total_unique_words > 0 else 0
            
            # Also consider if we have a good percentage of matches
            if len(matching_words) >= 2 and score > best_score:  # At least 2 words must match
                best_score = score
                best_match = orig_name
    
    # If we found a good match (at least 40% word overlap or 2+ words), use it
    if best_match and (best_score >= 0.4 or len(dept_words & set(re.findall(r'\b\w{3,}\b', normalize_for_matching(best_match)))) >= 2):
        return best_match
    
    return 'unknown'


def match_location_in_text(text: str, subcounties: dict, wards: dict, ward_to_subcounty: dict = None) -> tuple:
    """
    Try to match subcounty and ward names from the given text.
    Returns (subcounty_name, ward_name) where names are as found in database, or ('unknown', 'unknown').
    """
    if not text:
        return 'unknown', 'unknown'
    
    if not subcounties and not wards:
        # Database not connected, return unknown
        return 'unknown', 'unknown'
    
    # Expand shortcuts in text
    expanded_name = expand_shortcuts(text)
    normalized_expanded = normalize_for_matching(expanded_name)
    
    # Also try original without expansion
    normalized_original = normalize_for_matching(text)
    
    # Try to find matches - check from end of text (locations usually at the end)
    # Split by common delimiters and check each part
    parts = re.split(r'[,\-/\(\)]', expanded_name)
    parts_original = re.split(r'[,\-/\(\)]', text)
    
    # Also split by common words that might separate location info
    # Split on words like "in", "at", "for", "of" that might precede location
    location_separators = r'\b(in|at|for|of|within|across|throughout)\b'
    parts_expanded_sep = re.split(location_separators, expanded_name, flags=re.IGNORECASE)
    parts_original_sep = re.split(location_separators, text, flags=re.IGNORECASE)
    
    # Combine all parts, prioritizing expanded
    all_parts = parts + parts_original + parts_expanded_sep + parts_original_sep
    
    found_subcounty = None
    found_ward = None
    
    # First, try matching against the full normalized text strings
    # This catches cases like "Project CountyWide" where CountyWide is in the middle
    if subcounties:
        for norm_name, orig_name in subcounties.items():
            if not found_subcounty:
                # Check if subcounty name appears anywhere in the project name
                if norm_name in normalized_expanded or norm_name in normalized_original:
                    found_subcounty = orig_name
                elif normalized_expanded in norm_name or normalized_original in norm_name:
                    found_subcounty = orig_name
                else:
                    # Try word-by-word matching - check if all significant words of subcounty appear
                    subcounty_words = [w for w in norm_name.split() if len(w) > 2]
                    if len(subcounty_words) > 0:
                        all_words_found = all(
                            word in normalized_expanded or word in normalized_original 
                            for word in subcounty_words
                        )
                        if all_words_found:
                            found_subcounty = orig_name
    
    if wards:
        for norm_name, orig_name in wards.items():
            if not found_ward:
                if norm_name in normalized_expanded or norm_name in normalized_original:
                    found_ward = orig_name
                elif normalized_expanded in norm_name or normalized_original in norm_name:
                    found_ward = orig_name
                else:
                    ward_words = [w for w in norm_name.split() if len(w) > 2]
                    if len(ward_words) > 0:
                        all_words_found = all(
                            word in normalized_expanded or word in normalized_original 
                            for word in ward_words
                        )
                        if all_words_found:
                            found_ward = orig_name
    
    # Then check each part (from end to beginning, as locations are usually at the end)
    for part in reversed(all_parts):
        if not part or not part.strip():
            continue
        
        normalized_part = normalize_for_matching(part)
        if not normalized_part or len(normalized_part) < 3:  # Skip very short parts
            continue
        
        # Check if this part matches a subcounty (more flexible matching)
        if not found_subcounty and subcounties:
            for norm_name, orig_name in subcounties.items():
                # Try exact match first
                if normalized_part == norm_name:
                    found_subcounty = orig_name
                    break
                # Try substring match (either direction)
                if normalized_part in norm_name or norm_name in normalized_part:
                    # Prefer longer matches
                    if not found_subcounty or len(norm_name) > len(normalize_for_matching(found_subcounty or '')):
                        found_subcounty = orig_name
                # Try word-by-word matching for multi-word names
                part_words = set(normalized_part.split())
                name_words = set(norm_name.split())
                if len(part_words) > 0 and len(name_words) > 0:
                    # If most words match, consider it a match
                    common_words = part_words.intersection(name_words)
                    if len(common_words) >= min(2, len(name_words), len(part_words)):
                        found_subcounty = orig_name
        
        # Check if this part matches a ward (more flexible matching)
        if not found_ward and wards:
            for norm_name, orig_name in wards.items():
                # Try exact match first
                if normalized_part == norm_name:
                    found_ward = orig_name
                    break
                # Try substring match (either direction)
                if normalized_part in norm_name or norm_name in normalized_part:
                    # Prefer longer matches
                    if not found_ward or len(norm_name) > len(normalize_for_matching(found_ward or '')):
                        found_ward = orig_name
                # Try word-by-word matching for multi-word names
                part_words = set(normalized_part.split())
                name_words = set(norm_name.split())
                if len(part_words) > 0 and len(name_words) > 0:
                    # If most words match, consider it a match
                    common_words = part_words.intersection(name_words)
                    if len(common_words) >= min(2, len(name_words), len(part_words)):
                        found_ward = orig_name
        
        # If we found both, we can stop
        if found_subcounty and found_ward:
            break
    
    # Also try matching the full normalized project name (as fallback)
    if not found_subcounty and subcounties:
        for norm_name, orig_name in subcounties.items():
            # Check if subcounty name appears anywhere in project name
            if norm_name in normalized_expanded or norm_name in normalized_original:
                found_subcounty = orig_name
                break
            # Also check reverse (project name in subcounty name)
            if normalized_expanded in norm_name or normalized_original in norm_name:
                found_subcounty = orig_name
                break
    
    if not found_ward and wards:
        for norm_name, orig_name in wards.items():
            # Check if ward name appears anywhere in project name
            if norm_name in normalized_expanded or norm_name in normalized_original:
                found_ward = orig_name
                break
            # Also check reverse (project name in ward name)
            if normalized_expanded in norm_name or normalized_original in norm_name:
                found_ward = orig_name
                break
    
    # If we found a ward but not a subcounty, try to get subcounty from ward relationship
    if found_ward and found_ward != 'unknown' and not found_subcounty and ward_to_subcounty:
        if found_ward in ward_to_subcounty:
            found_subcounty = ward_to_subcounty[found_ward]
    
    return found_subcounty or 'unknown', found_ward or 'unknown'


def extract_projects_from_sheet(sheet, template_columns, subcounties, wards, departments, ward_to_subcounty):
    department, program = detect_program_and_department(sheet)
    header_row, header_values = detect_header_row(sheet)
    if not header_row:
        return []

    col_map = build_column_index_map(header_values)

    projects = []
    counter = 1
    max_row = sheet.max_row

    for row in range(header_row + 1, max_row + 1):
        first_val = clean_text(sheet.cell(row=row, column=1).value)
        # stop when we hit an empty first cell for a few rows in a row
        if not first_val:
            # allow occasional blank row; but break if we have already collected some
            if projects:
                empty_row = all(not clean_text(sheet.cell(row=row, column=c).value) for c in range(1, 5))
                if empty_row:
                    break
            continue

        row_data = {col: "" for col in template_columns}

        row_data["Department"] = department
        row_data["Program"] = program

        # Pull mapped columns
        for tmpl_col, col_idx in col_map.items():
            value = clean_text(sheet.cell(row=row, column=col_idx).value)
            row_data[tmpl_col] = value

        project_name = row_data.get("Project name and Location (Ward/Sub county/ county wide)", first_val)
        timeframe = row_data.get("Time frame", "")
        row_data["Project_ref"] = make_project_ref(project_name, program, timeframe, counter)
        
        # Match subcounty and ward from project name first, then try description if no matches
        project_description = row_data.get("Description of activities", "")
        subcounty, ward = match_location_in_text(project_name, subcounties, wards, ward_to_subcounty)
        
        # If no matches found in project name, try the description
        if (subcounty == 'unknown' or ward == 'unknown') and project_description:
            desc_subcounty, desc_ward = match_location_in_text(project_description, subcounties, wards, ward_to_subcounty)
            if subcounty == 'unknown' and desc_subcounty != 'unknown':
                subcounty = desc_subcounty
            if ward == 'unknown' and desc_ward != 'unknown':
                ward = desc_ward
                # If we found a ward from description but not subcounty, try to get it from ward relationship
                if subcounty == 'unknown' and ward != 'unknown' and ward_to_subcounty and ward in ward_to_subcounty:
                    subcounty = ward_to_subcounty[ward]
        
        # Use correct column names as per template
        row_data["Subcounty"] = subcounty
        row_data["ward"] = ward
        
        # Match department from detected department name
        db_department = match_department_name(department, departments)
        
        # If department is still unknown, try matching using the program name
        if db_department == 'unknown' and program:
            db_department = match_department_name(program, departments)
        
        row_data["db_department"] = db_department
        
        counter += 1

        # Status and Implementing Agency may not be present in ADP; leave blank for now.
        # They are already initialized as empty strings.

        projects.append(row_data)

    return projects


def main():
    if not ADP_PATH.exists():
        raise SystemExit(f"Source ADP workbook not found: {ADP_PATH}")
    if not TEMPLATE_PATH.exists():
        raise SystemExit(f"Template workbook not found: {TEMPLATE_PATH}")

    # Connect to database and fetch subcounties and wards
    print("Connecting to database...")
    connection = None
    try:
        connection = connect_to_database()
        print("Connected to database successfully.")
    except Exception as e:
        print(f"Warning: Could not connect to database: {e}")
        print("Will continue without database matching. Subcounty and Ward columns will be 'unknown'.")
        connection = None
    
    print("Fetching subcounties, wards, and departments from database...")
    subcounties, wards, ward_to_subcounty = fetch_subcounties_and_wards(connection)
    departments = fetch_departments(connection)
    print(f"Loaded {len(subcounties)} subcounties, {len(wards)} wards, and {len(departments)} departments from database.")
    print(f"Loaded {len(ward_to_subcounty)} ward-to-subcounty relationships.")
    
    if connection:
        connection.close()

    template_columns = load_template_columns()

    wb = openpyxl.load_workbook(ADP_PATH, data_only=True)

    all_rows = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_rows = extract_projects_from_sheet(sheet, template_columns, subcounties, wards, departments, ward_to_subcounty)
        print(f"{sheet_name}: extracted {len(sheet_rows)} project rows")
        all_rows.extend(sheet_rows)

    if not all_rows:
        print("No project rows detected; nothing to write.")
        return

    df = pd.DataFrame(all_rows, columns=template_columns)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUTPUT_PATH, index=False)
    print(f"Wrote {len(df)} rows to {OUTPUT_PATH}")
    
    # Print summary of matches
    print(f"\nMatching summary:")
    if 'Subcounty' in df.columns:
        unknown_subcounties = (df['Subcounty'].fillna('unknown') == 'unknown').sum()
        matched_subcounties = len(df) - unknown_subcounties
        print(f"  Subcounties matched: {matched_subcounties}/{len(df)}")
    
    if 'ward' in df.columns:
        unknown_wards = (df['ward'].fillna('unknown') == 'unknown').sum()
        matched_wards = len(df) - unknown_wards
        print(f"  Wards matched: {matched_wards}/{len(df)}")
    
    if 'db_department' in df.columns:
        unknown_departments = (df['db_department'].fillna('unknown') == 'unknown').sum()
        matched_departments = len(df) - unknown_departments
        print(f"  Departments matched: {matched_departments}/{len(df)}")


if __name__ == "__main__":
    main()


